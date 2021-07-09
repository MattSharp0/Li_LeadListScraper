from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, TimeoutException, ElementClickInterceptedException
from openpyxl import Workbook
from config import CREDENTIALS

# lead_list_name = input('Type lead list name: ')

# test lists:
# 3 Page, 54 lead list
lead_list_name = 'CaptialOne [mh]'
# 1 Page, 10 lead list
# lead_list_name = 'Ambassador [MH] [submitted]'

'''
Working on multipage support. 
Inconsistent detection of multipule pages and not detecting more than 2
Commented out write to excel funtion call
Set scrape to no longer run headless
'''


class scraper(webdriver.Chrome):
    '''
    Extends webdriver class with leadlist specifc methods
    '''

    def go_to_lead_list(self, credentials, lead_list_name):
        '''
        Scrapes LinkedIn profile links from provided sales navigator lead list and returns them as a list.

        :Args:
        list: lead_list_name

        :Returns:
        list of lead links

        Note:
        Requires sales navigator credentials imported as USERNAME and PASSWORD
        '''
        USERNAME, PASSWORD = credentials['USERNAME'], credentials['PASSWORD']

        # load sales navigator
        self.get('https://www.linkedin.com/sales')
        print('\n  Chrome driver object created\n    Signing into LinkedIn Sales Nav...')

        # Sign in
        try:
            username_box = self.find_element_by_id('username')
            username_box.send_keys(USERNAME)
            password_box = self.find_element_by_id('password')
            password_box.send_keys(PASSWORD)

            signin_button = self.find_element_by_xpath(
                '//*[@id="app__container"]/main/div[2]/form/div[3]/button')
            signin_button.click()
            print('\n  Signed in!\n    Loading homepage...')
        except NoSuchElementException as e:
            print(e)
            exit()
        except StaleElementReferenceException as e:
            print(e)
            exit()

        # Wait for homepage to load, then go to Lead lists
        try:
            WebDriverWait(self, 10).until(
                lambda b: b.find_element_by_id('ember14'))
            print('\n  Homepage loaded!')

            lead_lists = self.find_element_by_id('ember14')
            lead_lists.click()
            print('    Loading lead lists page...')
        except TimeoutException as e:
            print(e)
            exit(1)
        except StaleElementReferenceException as e:
            print(e)
            exit(1)
        except NoSuchElementException as e:
            print(e)
            exit(1)

        # Wait for Lead lists to load
        try:
            WebDriverWait(self, 10).until(lambda b: b.find_element_by_xpath(
                f"//*[text()='{lead_list_name}']//ancestor::a"))
            print('\n  Lead lists page loaded!')
        except TimeoutException as e:
            print(e)
            exit()

        # Go to target lead list
        try:
            self.find_element_by_xpath(
                f"//*[text()='{lead_list_name}']//ancestor::a").click()
            print(f'    Loading target lead list: {lead_list_name}...')
        except StaleElementReferenceException as e:
            print(
                f'\nError: Lead list: "{lead_list_name}" is no longer visible\n')
            print(e)
            exit()
        except NoSuchElementException as e:
            print(f'Error: Could not find list: "{lead_list_name}"')
            print(e)
            exit()

        return self

    def get_page_numbers(self):
        '''
        Finds current lead list page number and total number of pages

        :Args:
        Self: Chrome webdriver: Selenium driver on list page

        :Returns:
        int: current page number
        int: number of pages
        '''
        # Wait for page numbers to load
        WebDriverWait(self, 10).until(lambda b: b.find_elements_by_class_name(
            'artdeco-pagination__indicator--number'))

        # Find current page number
        current_page = self.find_element_by_xpath(
            '//*[@class="artdeco-pagination__indicator artdeco-pagination__indicator--number active selected ember-view"]/button/span[1]')
        current_page_number = current_page.text

        pages = len(self.find_elements_by_class_name(
            'artdeco-pagination__indicator--number'))

        return int(current_page_number), pages

    def create_list_of_links(self):
        '''
        Finds and adds profile links to list

        :Args:
        Self: driver: Selenium driver on list page

        :Returns:
        current_page_links - a list of profile links from page
        '''

        WebDriverWait(self, 10).until(lambda b: b.find_elements_by_class_name(
            'lists-detail__view-profile-name-link'))
        print('\n  Target lead list page loaded!')

        print('    Scraping lead profile links...')
        profile_links = self.find_elements_by_class_name(
            'lists-detail__view-profile-name-link')

        current_page_links = [profile_link.get_attribute(
            'href') for profile_link in profile_links]

        print(f'\n  Saved {len(profile_links)} links from page to list!')
        return current_page_links


def write_to_excel(link_list, lead_list_name):
    '''
    Takes a list of links and writes them to an excel document with the list name as the title

    Requires list of strings/ints

    Requires list name
    '''

    # Quit if list has no items
    if len(link_list) < 1:
        print('\nList empty!')
        exit(1)

    # Create document and sheet
    print('\nWriting list to excel...')
    wb = Workbook()
    sheet = wb.create_sheet('lead list', 0)

    # Format title
    title = ((lead_list_name.split()[0]) + ' leads')

    # Write title
    sheet['A1'].value = title

    # Write links to A column
    r = 2
    for link in link_list:
        box = sheet.cell(r, 1)
        box.value = link
        r += 1

    # Save document with title
    wb.save(f'{title}.xlsx')

    print(f'\nSaved {len(link_list)} lead links to {title}.xlsx')


options = Options()
# options.add_argument('--headless')
# options.add_argument('window-size=1920x1080')

# create driver object
with scraper(options=options) as browser:
    browser.go_to_lead_list(CREDENTIALS, lead_list_name)

    list_of_lead_links = browser.create_list_of_links()

    current_page, pages = browser.get_page_numbers()

    # While multiple pages exist, go page by page and copy lead links to list
    while current_page < pages:
        # Find and click next page button
        try:
            WebDriverWait(browser, 10).until(lambda b: b.find_elements_by_class_name(
                'artdeco-pagination__button--next'))

            next_page = browser.find_element_by_class_name(
                'artdeco-pagination__button--next')
            next_page.click()
        except ElementClickInterceptedException as e:
            print(e)
            exit(1)

        # Update current page number
        current_page, pages = browser.get_page_numbers()
        print(f'\n    Loading page {current_page}/{pages}...')

        # For debugging:
        list_of_lead_links.append(f'{current_page}/{pages}')

        # Add links to main list
        for link in browser.create_list_of_links():
            list_of_lead_links.append(link)

    # check list validity and exit
    if len(list_of_lead_links) > 1:
        print(
            f'\nScrape complete!\n{len(list_of_lead_links)} links added to list')
        write_to_excel(lead_list_name, list_of_lead_links)
    else:
        print('\nError! List is empty')
        exit(1)
