from openpyxl import Workbook


def write_to_excel(link_list, lead_list_name):
    '''
    Takes a list of links and writes them to an excel document with the list name as the title

    :Params:
    List: link_list: list of strings
    Str: lead_list_name: name of lead list
    '''

    # Quit if list has no items
    if len(link_list) < 1:
        print('\nList empty!')
        exit(1)

    # Create document and sheet
    print('\nWriting list to excel...')
    wb = Workbook()
    sheet = wb.create_sheet('lead list', 0)

    # Format title
    title = ((lead_list_name.split()[0]) + ' leads')

    # Write title
    sheet['A1'].value = title

    # Write links to A column
    r = 2
    for link in link_list:
        box = sheet.cell(r, 1)
        box.value = link
        r += 1

    # Save document with title
    wb.save(f'{title}.xlsx')

    print(f'\nSaved {len(link_list)} lead links to {title}.xlsx')
